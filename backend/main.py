from __future__ import annotations

import io
import math
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

app = FastAPI(title="Dynamic Excel Builder API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MAX_UPLOAD_BYTES = 100 * 1024 * 1024


def safe_filename(name: str, default: str = "export") -> str:
    name = (name or default).strip()
    name = re.sub(r"[^\w\-.\u0080-\uffff ]+", "_", name, flags=re.UNICODE)
    return name or default


def unique_columns(columns: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for i, col in enumerate(columns):
        base = str(col).strip() if col is not None and not (isinstance(col, float) and math.isnan(col)) else ""
        if not base or base.lower().startswith("unnamed:"):
            base = f"Column_{i + 1}"
        count = seen.get(base, 0)
        result.append(base if count == 0 else f"{base}_{count}")
        seen[base] = count + 1
    return result


def json_safe(value: Any) -> Any:
    """Convert Excel/Pandas values to clean JSON values.

    Important: Excel dates often arrive as pandas Timestamp/datetime values.
    If the time part is empty/midnight, return only YYYY-MM-DD so the sub sheet
    does not show unwanted timestamps like 2026-06-01T00:00:00.
    """
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    df.columns = unique_columns(list(df.columns))
    return df


def excel_date_display(value: Any, number_format: str = "") -> str:
    """Return the cell value as a clean Excel-visible date string.

    Main rule: never send ISO datetime text such as 2022-05-19T00:00:00 to
    the frontend.  If Excel stores a date with midnight time, only the date is
    returned.  The date order/separator follows the cell's Excel number format
    as closely as possible.
    """
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if not isinstance(value, (datetime, date)):
        return str(value)

    dt = value if isinstance(value, datetime) else datetime(value.year, value.month, value.day)
    fmt_original = number_format or ""
    fmt = fmt_original.lower()
    fmt = re.sub(r"\[[^\]]+\]", "", fmt)
    fmt = fmt.split(";")[0].strip()

    has_real_time = bool(dt.hour or dt.minute or dt.second or dt.microsecond)
    has_time_format = any(token in fmt for token in ["h", "s", "am/pm"])

    # If actual time exists in Excel, keep it, but never use ISO/T format.
    if has_real_time and has_time_format:
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    # Common Excel date formats are converted to a plain visible string.
    if not any(ch in fmt for ch in ["d", "m", "y"]):
        return dt.strftime("%Y-%m-%d")

    sep = "/" if "/" in fmt else "-" if "-" in fmt else "." if "." in fmt else "-"
    d_pos = fmt.find("d") if "d" in fmt else 999
    m_pos = fmt.find("m") if "m" in fmt else 999
    y_pos = fmt.find("y") if "y" in fmt else 999
    order = [p for p, pos in sorted([("d", d_pos), ("m", m_pos), ("y", y_pos)], key=lambda x: x[1]) if pos != 999]
    if not order:
        order = ["y", "m", "d"]

    d_token = re.search(r"d+", fmt)
    m_token = re.search(r"m+", fmt)
    y_token = re.search(r"y+", fmt)

    day = f"{dt.day:02d}" if d_token and len(d_token.group(0)) >= 2 else str(dt.day)
    month_token_len = len(m_token.group(0)) if m_token else 2
    if month_token_len >= 4:
        month = dt.strftime("%B")
    elif month_token_len == 3:
        month = dt.strftime("%b")
    else:
        month = f"{dt.month:02d}" if month_token_len >= 2 else str(dt.month)
    year = f"{dt.year % 100:02d}" if y_token and len(y_token.group(0)) <= 2 else str(dt.year)

    parts_map = {"d": day, "m": month, "y": year}
    return sep.join(parts_map[p] for p in order)


def excel_cell_display_value(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return excel_date_display(value, getattr(cell, "number_format", ""))
    return json_safe(value)


def read_xlsx_with_display_values(content: bytes, header_row: int, formulas: str = "values") -> pd.DataFrame:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=(formulas == "values"), read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg:
            raise HTTPException(status_code=400, detail="Password-protected files are not supported.")
        raise HTTPException(status_code=400, detail="Unable to read this Excel file. It may be corrupted or unsupported.")

    header_excel_row = max(header_row, 0) + 1
    raw_headers = []
    for cell in ws[header_excel_row]:
        raw_headers.append(excel_cell_display_value(cell))
    columns = unique_columns(raw_headers)

    data_rows: List[List[Any]] = []
    for excel_row in ws.iter_rows(min_row=header_excel_row + 1, max_col=len(columns)):
        row_values = [excel_cell_display_value(cell) for cell in excel_row]
        if any(str(v).strip() for v in row_values):
            data_rows.append(row_values)

    if not columns or not any(str(c).strip() for c in columns):
        raise HTTPException(status_code=400, detail="No data found.")

    df = pd.DataFrame(data_rows, columns=columns)
    df = df.replace("", pd.NA).dropna(how="all").dropna(axis=1, how="all").fillna("")
    df.columns = unique_columns(list(df.columns))
    return df


def read_excel_upload(file: UploadFile, content: bytes, header_row: int, formulas: str = "values") -> pd.DataFrame:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload a valid Excel file.")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="No data found.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large. Maximum supported upload is 100 MB.")
    if suffix == ".xlsx":
        df = read_xlsx_with_display_values(content, header_row, formulas)
        if df.empty or len(df.columns) == 0:
            raise HTTPException(status_code=400, detail="No data found.")
        return df
    try:
        # .xls fallback through pandas/xlrd. Macro/script execution never occurs.
        df = pd.read_excel(io.BytesIO(content), header=header_row, dtype=object)
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg:
            raise HTTPException(status_code=400, detail="Password-protected files are not supported.")
        raise HTTPException(status_code=400, detail="Unable to read this Excel file. It may be corrupted or unsupported.")
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg or "file is not a zip" in msg:
            raise HTTPException(status_code=400, detail="Password-protected or corrupted files are not supported.")
        raise HTTPException(status_code=400, detail=f"Unable to process Excel file: {exc}")
    df = normalize_dataframe(df)
    if df.empty or len(df.columns) == 0:
        raise HTTPException(status_code=400, detail="No data found.")
    return df


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/api/parse-excel")
async def parse_excel(
    file: UploadFile = File(...),
    header_row: int = Form(0),
    formulas: str = Form("values"),
) -> Dict[str, Any]:
    content = await file.read()
    df = read_excel_upload(file, content, max(header_row, 0), formulas)
    rows = [{col: json_safe(row[col]) for col in df.columns} for _, row in df.iterrows()]
    columns = [
        {
            "id": re.sub(r"\W+", "_", str(col).strip().lower(), flags=re.UNICODE).strip("_") or f"column_{idx+1}",
            "originalName": col,
            "displayName": col,
            "dataType": str(df[col].dropna().map(type).iloc[0].__name__) if not df[col].dropna().empty else "text",
        }
        for idx, col in enumerate(df.columns)
    ]
    return {
        "fileName": file.filename,
        "sheetName": "Sheet1",
        "headerRowIndex": header_row,
        "columns": columns,
        "rows": rows,
        "rowCount": len(rows),
        "columnCount": len(columns),
    }


@app.post("/api/import-working-sheet")
async def import_working_sheet(file: UploadFile = File(...), header_row: int = Form(0)) -> Dict[str, Any]:
    content = await file.read()
    df = read_excel_upload(file, content, max(header_row, 0), "values")
    rows = [[json_safe(row[col]) for col in df.columns] for _, row in df.iterrows()]
    return {"columns": [str(c) for c in df.columns], "rows": rows, "rowCount": len(rows)}


class ExportColumn(BaseModel):
    id: str
    displayName: str
    sourceColumn: Optional[str] = None
    isCustom: bool = False


class ExportPayload(BaseModel):
    filename: str = "export"
    columns: List[ExportColumn]
    rows: List[Dict[str, Any]] = Field(default_factory=list)
    # Frontend sends keys like "0:column_id" with CSS-like style values.
    cellStyles: Dict[str, Dict[str, Any]] = Field(default_factory=dict)
    # Frontend sends current live-grid sizes so exported Excel opens like the live sheet.
    columnWidths: List[float] = Field(default_factory=list)  # pixels from Handsontable
    rowHeights: List[float] = Field(default_factory=list)    # pixels from Handsontable


def row_value(row: Dict[str, Any], col_id: str) -> Any:
    val = row.get(col_id, "")
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return str(val)
    # Ensure exported Excel never receives browser ISO timestamp strings.
    if isinstance(val, str):
        m = re.match(r"^(\d{4}-\d{2}-\d{2})T00:00:00(?:\.000)?(?:Z)?$", val)
        if m:
            return m.group(1)
    return val


def css_hex_to_argb(value: Any) -> Optional[str]:
    if not value:
        return None
    text = str(value).strip()
    if re.match(r"^#[0-9a-fA-F]{6}$", text):
        return "FF" + text[1:].upper()
    return None


def css_font_size_to_points(value: Any) -> Optional[float]:
    if not value:
        return None
    text = str(value).strip().lower()
    try:
        if text.endswith("px"):
            return round(float(text[:-2]) * 0.75, 2)
        if text.endswith("pt"):
            return float(text[:-2])
        return float(text)
    except Exception:
        return None


THIN_EXCEL_BORDER = Border(
    left=Side(style="thin", color="D9DDE5"),
    right=Side(style="thin", color="D9DDE5"),
    top=Side(style="thin", color="D9DDE5"),
    bottom=Side(style="thin", color="D9DDE5"),
)


def pixels_to_excel_width(px: Any) -> Optional[float]:
    try:
        value = float(px)
        if value <= 0:
            return None
        # Excel column width approximates character count, not pixels.
        return round(max(8, min(80, (value - 5) / 7)), 2)
    except Exception:
        return None


def pixels_to_points(px: Any) -> Optional[float]:
    try:
        value = float(px)
        if value <= 0:
            return None
        return round(max(15, min(240, value * 0.75)), 2)
    except Exception:
        return None


MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def normalize_two_digit_year(year_text: str) -> int:
    year = int(year_text)
    if len(year_text) == 2:
        return 2000 + year if year <= 49 else 1900 + year
    return year


def parse_date_string_for_excel(value: Any) -> Optional[tuple[datetime, str]]:
    """Convert text-looking dates into real Excel dates to avoid Excel warnings.

    Example warning fixed: "Text Date with 2-Digit Year" for values like 1-Apr-26.
    We export these as real date cells with a 4-digit year format.
    """
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None

    # 1-Apr-26 / 01-Apr-2026 / 1-April-26
    m = re.match(r"^(\d{1,2})[-\s]([A-Za-z]{3,9})[-\s](\d{2}|\d{4})$", text)
    if m:
        day = int(m.group(1))
        month = MONTH_MAP.get(m.group(2).lower())
        year_text = m.group(3)
        year = normalize_two_digit_year(year_text)
        if month:
            # Store as a real Excel date but keep the same visible style.
            return datetime(year, month, day), ("d-mmm-yy" if len(year_text) == 2 else "d-mmm-yyyy")

    # dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy. Indian/user expected format = day first.
    m = re.match(r"^(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2}|\d{4})$", text)
    if m:
        day = int(m.group(1))
        month = int(m.group(2))
        year_text = m.group(3)
        year = normalize_two_digit_year(year_text)
        if 1 <= day <= 31 and 1 <= month <= 12:
            return datetime(year, month, day), ("dd/mm/yy" if len(year_text) == 2 else "dd/mm/yyyy")

    # yyyy-mm-dd / yyyy/mm/dd
    m = re.match(r"^(\d{4})[\/\-.](\d{1,2})[\/\-.](\d{1,2})$", text)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        day = int(m.group(3))
        if 1 <= day <= 31 and 1 <= month <= 12:
            return datetime(year, month, day), "yyyy-mm-dd"

    return None


def apply_excel_cell_style(cell: Any, style: Dict[str, Any]) -> None:
    color = css_hex_to_argb(style.get("color")) if style else None
    fill = css_hex_to_argb(style.get("backgroundColor")) if style else None
    size = css_font_size_to_points(style.get("fontSize")) if style else None
    cell.font = Font(
        name=(style or {}).get("fontFamily") or "Calibri",
        bold=str((style or {}).get("fontWeight", "")) in {"700", "bold"},
        italic=str((style or {}).get("fontStyle", "")) == "italic",
        color=color,
        size=size or 11,
    )
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    css_vertical = (style or {}).get("verticalAlign") or "middle"
    excel_vertical = "center" if css_vertical == "middle" else css_vertical
    cell.alignment = Alignment(
        wrap_text=True,
        vertical=excel_vertical,
        horizontal=(style or {}).get("textAlign") or "left",
    )
    cell.border = THIN_EXCEL_BORDER


def remove_excel_filter_and_table_xml(xlsx_bytes: bytes) -> bytes:
    """Hard-remove Excel table/filter XML from generated XLSX.

    This is intentionally strict because Excel can display green table headers and
    dropdown filter arrows if any <tableParts>, <autoFilter>, table relationships,
    or table content types remain in the XLSX package. The user wants a plain
    worksheet header only.
    """
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", spreadsheet_ns)
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")

    def remove_nodes(root: ET.Element, local_name: str) -> None:
        for parent in list(root.iter()):
            for child in list(parent):
                if child.tag.endswith("}" + local_name) or child.tag == local_name:
                    parent.remove(child)

    input_io = io.BytesIO(xlsx_bytes)
    output_io = io.BytesIO()

    with zipfile.ZipFile(input_io, "r") as zin, zipfile.ZipFile(output_io, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename

            # Remove all Excel table definition files completely.
            if name.startswith("xl/tables/"):
                continue

            data = zin.read(name)

            # Remove autoFilter/tableParts from worksheets.
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                try:
                    root = ET.fromstring(data)
                    remove_nodes(root, "autoFilter")
                    remove_nodes(root, "tableParts")
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass

            # Remove relationships pointing to table XML.
            elif name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
                try:
                    root = ET.fromstring(data)
                    for rel in list(root):
                        target = rel.attrib.get("Target", "")
                        rel_type = rel.attrib.get("Type", "")
                        if "table" in target.lower() or rel_type.endswith("/table"):
                            root.remove(rel)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass

            # Remove table content type declarations.
            elif name == "[Content_Types].xml":
                try:
                    root = ET.fromstring(data)
                    for child in list(root):
                        part_name = child.attrib.get("PartName", "")
                        content_type = child.attrib.get("ContentType", "")
                        if part_name.startswith("/xl/tables/") or content_type.endswith("spreadsheetml.table+xml"):
                            root.remove(child)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass

            # Remove filter database defined names if any exist.
            elif name == "xl/workbook.xml":
                try:
                    root = ET.fromstring(data)
                    for parent in list(root.iter()):
                        for child in list(parent):
                            text = "".join(child.itertext()) if list(child) or child.text else (child.text or "")
                            if "_FilterDatabase" in child.attrib.get("name", "") or "_FilterDatabase" in text:
                                parent.remove(child)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass

            zout.writestr(item, data)

    return output_io.getvalue()


@app.post("/api/export/excel")
def export_excel(payload: ExportPayload) -> Response:
    """Export a plain worksheet only.

    IMPORTANT FIX:
    This endpoint intentionally uses XlsxWriter and never creates Excel Tables,
    never enables AutoFilter, and never freezes panes. Therefore Excel cannot
    display the green table header or filter dropdown icons shown in the user's
    screenshot.
    """
    import xlsxwriter

    stream = io.BytesIO()
    workbook = xlsxwriter.Workbook(stream, {"in_memory": True, "strings_to_urls": False})
    worksheet = workbook.add_worksheet("Working Sheet")

    # Show normal Excel gridlines and keep worksheet plain.
    worksheet.hide_gridlines(0)

    format_cache: Dict[str, Any] = {}

    def xlsx_color(hex_value: Any) -> Optional[str]:
        if not hex_value:
            return None
        text = str(hex_value).strip()
        return text if re.match(r"^#[0-9a-fA-F]{6}$", text) else None

    def css_size_to_points(value: Any) -> int:
        pts = css_font_size_to_points(value)
        return int(round(pts)) if pts else 11

    def make_format(style: Optional[Dict[str, Any]] = None, *, header: bool = False, date_format: Optional[str] = None):
        style = style or {}
        key = str(sorted(style.items())) + f"|header={header}|date={date_format or ''}"
        if key in format_cache:
            return format_cache[key]

        props: Dict[str, Any] = {
            "font_name": style.get("fontFamily") or "Calibri",
            "font_size": css_size_to_points(style.get("fontSize")),
            "text_wrap": True,
            "valign": "vcenter" if (style.get("verticalAlign") or "middle") == "middle" else (style.get("verticalAlign") or "vcenter"),
            "align": style.get("textAlign") or "left",
        }

        if header:
            # Plain header only: bold black text, no fill, no filter, no table.
            props.update({"bold": True, "font_color": "#000000", "bg_color": None, "pattern": 0})

        if str(style.get("fontWeight", "")) in {"700", "bold"}:
            props["bold"] = True
        if str(style.get("fontStyle", "")) == "italic":
            props["italic"] = True
        if xlsx_color(style.get("color")):
            props["font_color"] = xlsx_color(style.get("color"))
        if xlsx_color(style.get("backgroundColor")):
            props["bg_color"] = xlsx_color(style.get("backgroundColor"))
        if date_format:
            props["num_format"] = date_format

        # No border on header; very light border on data only if styled by user is not required.
        fmt = workbook.add_format({k: v for k, v in props.items() if v is not None})
        format_cache[key] = fmt
        return fmt

    header_format = make_format(header=True)

    # Write header as a normal row. No worksheet.autofilter(), no add_table().
    for col_idx, col in enumerate(payload.columns):
        worksheet.write(0, col_idx, col.displayName, header_format)

    # Write body exactly from live sheet. Date-looking text is converted to real
    # Excel date values to avoid green triangle/date text warnings.
    for row_idx, row in enumerate(payload.rows, start=1):
        for col_idx, col in enumerate(payload.columns):
            value = row_value(row, col.id)
            style = payload.cellStyles.get(f"{row_idx - 1}:{col.id}", {})
            parsed_date = parse_date_string_for_excel(value)
            if parsed_date:
                date_value, date_fmt = parsed_date
                worksheet.write_datetime(row_idx, col_idx, date_value, make_format(style, date_format=date_fmt))
            else:
                worksheet.write(row_idx, col_idx, value, make_format(style))

    # Match live sheet widths/heights where possible.
    for col_idx, col in enumerate(payload.columns):
        live_width = payload.columnWidths[col_idx] if col_idx < len(payload.columnWidths) else None
        converted_width = pixels_to_excel_width(live_width)
        if not converted_width:
            max_len = len(str(col.displayName))
            for row in payload.rows:
                max_len = max(max_len, len(str(row_value(row, col.id))))
            converted_width = min(max(max_len + 2, 10), 60)
        worksheet.set_column(col_idx, col_idx, converted_width)

    worksheet.set_row(0, 22, header_format)
    for row_idx in range(len(payload.rows)):
        live_height = payload.rowHeights[row_idx] if row_idx < len(payload.rowHeights) else None
        worksheet.set_row(row_idx + 1, pixels_to_points(live_height) or 21)

    workbook.close()
    clean_xlsx = remove_excel_filter_and_table_xml(stream.getvalue())

    name = safe_filename(payload.filename)
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return Response(
        clean_xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


def register_unicode_font() -> str:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("AppUnicode", path))
                return "AppUnicode"
            except Exception:
                continue
    return "Helvetica"


@app.post("/api/export/pdf")
def export_pdf(payload: ExportPayload) -> Response:
    font_name = register_unicode_font()
    stream = io.BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "header",
        parent=cell_style,
        fontSize=7,
        leading=9,
    )

    def para(value: Any, style: ParagraphStyle) -> Paragraph:
        text = "" if value is None else str(value)
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(text, style)

    data: List[List[Any]] = [[para(c.displayName, header_style) for c in payload.columns]]
    for row in payload.rows:
        data.append([para(row_value(row, c.id), cell_style) for c in payload.columns])

    page_width = landscape(A4)[0] - 16 * mm
    ncols = max(len(payload.columns), 1)
    col_width = page_width / ncols
    widths = [col_width] * ncols

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    doc.build([table])
    name = safe_filename(payload.filename)
    if not name.lower().endswith(".pdf"):
        name += ".pdf"
    return Response(
        stream.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("PORT", "8765"))
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="info")
